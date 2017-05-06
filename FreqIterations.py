import re
from openpyxl import load_workbook

def ReadValues(file, sheetname, cellStr, cellStart, cellEnd):
        values = []

        for i in range(cellStart, cellEnd + 1):
            cellIndex = cellStr
            cellIndex+= str(i)
            values.append((sheetname[cellIndex].value))
        return values

def RecommendChecker(totalGU):
    if(totalGU > 500):
        return 1
    else:
        return 0

def PickMaxGU():
    return guValues.index(max(guValues))

def PickGUCellIndex(maxGUIndex):
    return(guCellStart + (maxGUIndex))

def PickParamsForGU(guCell):
    cellStr = 'B'
    cellStr+= str(guCell)
    return relevantParameterDict[cellStr]

def PickMaxImpactParam(list, iterationIndex):
    for element in list:
        frequencyCell = 'G'  ##Note: For the corresponding param
        cellIndex = str(element + 1)
        frequencyCell += cellIndex

        ### Writing impact by reducing frequency.
        if (iterationIndex == 0):
            WB1 = load_workbook(fileName)
            sheet1 = WB1[defSheetName]
            originalVal = float(sheet1[frequencyCell].value)
            updateVal = (originalVal * 0.9)
            sheet1[frequencyCell].value = updateVal
            WB1.save(fileName)
            iterationIndex += 1

        else:
            ## Reading the impact and for further iterations.
            WB2 = load_workbook(fileName)
            sheet2 = WB2[defSheetName]
            guValuesNow = ReadValues(WB2, sheet2, guCellIndex, guCellStart, guCellEnd)

            ## Verification whether the cell accessed is a formula related cell.
            for val in guValuesNow:
                if(type(val) is str):
                    eachValList = re.findall(r"[\w']+", val)
                    sheetValListIndices = []
                    cellValListIndices = []
                    index = 0
                    print(eachValList)
                    for eachVal in eachValList:
                        sheetFlag = False
                        for sheet in sheetNames:
                            if(eachVal == sheet):
                                print("Its a sheet")
                                sheetFlag = True
                                sheetValListIndices.append(index)
                        if(sheetFlag is False):
                            print("Its a cell")
                            cellValListIndices.append(index)
                        index+= 1
                    print(sheetValListIndices)
                    print(cellValListIndices)
                    break

            ## Access strings if it is a formula cell.





        ## Making it back (i.e. Clearing the impact)
        #sheet[frequencyCell].value = originalVal

def ReduceEmissions():
    maxGUIndex = PickMaxGU()
    guCellIndex = PickGUCellIndex(maxGUIndex)
    paramsList = PickParamsForGU(guCellIndex)
    reductionTimes = 0
    PickMaxImpactParam(paramsList, reductionTimes)

########### Parameters set up - Hard coded values ##################
fileName = 'BetaCalculator_Update 3.0.3.xlsx'
defSheetName = 'Data Aggregation' #Default sheet name.
sheetNames = ['Start', 'Household Input', 'Style Change', 'GreenUser', 'Data Aggregation', 'Report Preferences',
              'Food Footprint Report', 'Diet Calculator', 'Database Food', 'total footprint report',
              'export alternative lifestyles', 'Diet Input', 'household calculator', 'back end data',
              'GreenStyleUser', 'explore alternative diets', 'Sub_Data', 'Transport Info', 'Dropdown',
              'Electricity mix footprint', 'old data sheet', 'useful links']


WB = load_workbook(fileName, data_only= True)
sheet = WB[defSheetName]

freqCellIndex = 'G'
freqCellStart = 2
freqCellEnd = 17

guCellIndex = 'B'
guCellStart = 34
guCellEnd = 38  ##Please note: GU value is not considered here.

frequency = ReadValues(fileName, sheet, freqCellIndex, freqCellStart, freqCellEnd)
guValues = ReadValues(fileName, sheet, guCellIndex, guCellStart, guCellEnd)
guTotal = sum(guValues)

### Change it to read from excel once the format has been changed.
### Todo: Replace string reading using re module.
relevantParameterDict = {
    'B34' : [16],
    'B35' : [12,13,14,15],
    'B36' : [1,2,3,4,5,6,9,10,11],
    'B37' : [7,8]
}

recomFlag = RecommendChecker(guTotal)

if(recomFlag):
    ReduceEmissions()
